from django.shortcuts import render, redirect
from django.http import JsonResponse
from rest_framework.views import APIView
from rest_framework.response import Response


from django.conf import settings

import os  
import requests
import openpyxl
from pathlib import Path

def process_excel(file, row_num):
    wb_obj = openpyxl.load_workbook(file)
    sheet = wb_obj.active
    r = {}
    c = 0 
    for row in sheet.iter_rows():
        # if row == row_num :
        c+=1
        if c == row_num :
            
            for cell in row:
                if cell.value:
                    r[ f'Column{cell.column_letter}' ] = cell.value
            # print(r)
            return r
           
    return {}

def analays_excel(file, row_num, num_rec):
    wb_obj = openpyxl.load_workbook(file)
    sheet = wb_obj.active
    r = {}
    c = 0 
    for row in sheet.iter_rows():
        # if row == row_num :
        c+=1
        if c == row_num :
            
            for cell in row:
                if cell.value:
                    r[ f'Column{cell.column_letter}' ] = cell.value
            
            break
    c = 0 
    checked = 0
    l = [ {"string" : 0, "integer" : 0, "decimal" : 0} for i in r ]
    for row in sheet.iter_rows():
        c += 1
        
        if c <= row_num :
            continue

        if checked == num_rec:
            break
        checked += 1 
        
        rc = 0
        for cell in row:
            
            if cell.value:
                if str(cell.value).isdigit() :
                    l[rc]["integer"] += 1
                    
                elif str(cell.value).replace(',','').isdecimal() :
                    
                    l[rc]['decimal'] += 1
                
                else:
                    l[rc]["string"] += 1 
                    
                rc += 1
                
                
    types = [] 
    for i in l:
        max_ = ""
        max_n = 0
        for j in i :
            if max_n <= i[j] :
                max_ = j
                max_n = i[j]
                
        types.append(max_)
    
    return { i:j for i,j in zip(r,types) }

            
    

def download_file(url):

    # this will grab the filename from the url
    filename = url.split('/')[-1]

    # print(f'Downloading {filename}')

    r = requests.get(url)
    path = os.path.join(settings.BASE_DIR, filename)
    with open(path, 'wb') as output_file:
        output_file.write(r.content)
        
    return path
    
    

def redirect_home(request):
    return redirect("readXLSheader")



def analysis(request):
    
    if request.method == "GET" :
        return render(request, "indext3.html")
    else:
        url = request.POST.get("link")
        row_req = 1
        
        if request.POST.get("row", None):

            row_req = int(request.POST.get("row"))
        numrec = int(request.POST.get("numrec", 0))
        
        file = download_file(url)
        
        print(file)
        
        col = process_excel(file, row_req)
        
        analysed = analays_excel(file, row_req, numrec)
        
    return JsonResponse( {  "ColumnVales" : col , "types" : analysed })
from datetime import datetime
import mysql.connector as mysql
import json
def log_to_db(data):
    host = settings.DB_HOST
    user = settings.DB_USER
    password = settings.DB_PASS
    db = settings.DB
    db_con = mysql.connect(host=host, user=user, database=db, password=password)
    query = "INSERT INTO XLSHeaderLog_100 (User, EventTime, XLSLink, JsonResult)  VALUES (%s,%s,%s,%s)"
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    user = "1"
    link = data['link']
    res = json.dumps(data['res'])
    val = (user, timestamp, link, res)
    
    cursor = db_con.cursor()
    
    cursor.execute(query, val) 
    db_con.commit()
    
    return cursor.rowcount


from .serializers import ExcelSerializer,ExcelAnalaysisSerializer
from rest_framework import status
class Task1(APIView):
    def get(self, request,):
        print(request.data)
        return Response({"data" : "Provide the XML data and use POST request"})
    
    def post(self, request, format=None):
        print(request.data)
        serializer = ExcelSerializer(data=request.data)
        if serializer.is_valid():
            link = request.data.get("file", None)
            if not link:
                return Response({"error" : "file is required"}) 
            
            row = request.data.get("row", 1)
            
            file = download_file(link)
            print(file)
            col = process_excel(file, row)
            
            return Response(col, status=status.HTTP_200_OK)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class Task2(APIView):
    def get(self, request,):
        print(request.data)
        return Response({"data" : "Provide the XML data and use POST request"})
    
    def post(self, request, format=None):
        print("TASK 2 ---------------")
        serializer = ExcelSerializer(data=request.data)
        if serializer.is_valid():
            link = request.data.get("file", None)
            if not link:
                return Response({"error" : "file is required"}) 
            
            row = request.data.get("row", 1)
            
            file = download_file(link)
            # print(file)
            col = process_excel(file, row)
            print("Data Inserted ", log_to_db({"link" : link, "res" : col}))
            return Response(col, status=status.HTTP_200_OK)
        else:
            print("INVALID")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class Task3(APIView):
    def get(self, request,):
        return Response({"data" : "Provide the XML data(file, row, row_number) and use POST request"})
    
    def post(self, request, format=None):
        print(request.data)
        serializer = ExcelAnalaysisSerializer(data=request.data)
        if serializer.is_valid():
            link = request.data.get("file", None)
            if not link:
                return Response({"error" : "file is required"}) 
            
            row = request.data.get("row", 1)
            numrec = int(request.data.get("row_number", 0))
            
            file = download_file(link)
            print(file)
            col = process_excel(file, row)
        
            analysed = analays_excel(file, row, numrec)
            return Response({  "ColumnVales" : col , "types" : analysed }, status=status.HTTP_200_OK)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)